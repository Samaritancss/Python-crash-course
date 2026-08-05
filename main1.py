import openpyxl

inv_file = openpyxl.load_workbook("inventory.xlsx")
inventory_id = inv_file["Inventory ID"]

reorder_level = {}

print(inventory_id.max_row)

for inventory_row in range(5, inventory_id.max_row + 1):
    quantity_in_stock = inventory_id.cell(inventory_row, 4)

    if quantity_in_stock in reorder_level:
        current_inventory = reorder_level[quantity_in_stock]
        reorder_level[quantity_in_stock] = current_inventory + 1
    else:
        print("adding a new reorder level")
        reorder_level[quantity_in_stock] = 1






